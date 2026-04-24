# import required modules
import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox

# file path
file_path = 'C:\\Users\\Admin\\Desktop\\excel.xlsx'

# check if file exists, else create it
if os.path.exists(file_path):
    wb = load_workbook(file_path)
else:
    wb = Workbook()
    wb.save(file_path)

# select active sheet
sheet = wb.active


# function to setup excel sheet
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 40

    # headers (only once)
    if sheet.max_row == 1:
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Course"
        sheet.cell(row=1, column=3).value = "Semester"
        sheet.cell(row=1, column=4).value = "Form Number"
        sheet.cell(row=1, column=5).value = "Contact Number"
        sheet.cell(row=1, column=6).value = "Email ID"
        sheet.cell(row=1, column=7).value = "Address"


# focus functions
def focus1(event): course_field.focus_set()
def focus2(event): sem_field.focus_set()
def focus3(event): form_no_field.focus_set()
def focus4(event): contact_no_field.focus_set()
def focus5(event): email_id_field.focus_set()
def focus6(event): address_field.focus_set()


# clear all fields
def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


# insert data into excel
def insert():
    if (name_field.get() == "" or
        course_field.get() == "" or
        sem_field.get() == "" or
        form_no_field.get() == "" or
        contact_no_field.get() == "" or
        email_id_field.get() == "" or
        address_field.get() == ""):

        messagebox.showwarning("Input Error", "Please fill all fields")
        return

    current_row = sheet.max_row + 1

    sheet.cell(row=current_row, column=1).value = name_field.get()
    sheet.cell(row=current_row, column=2).value = course_field.get()
    sheet.cell(row=current_row, column=3).value = sem_field.get()
    sheet.cell(row=current_row, column=4).value = form_no_field.get()
    sheet.cell(row=current_row, column=5).value = contact_no_field.get()
    sheet.cell(row=current_row, column=6).value = email_id_field.get()
    sheet.cell(row=current_row, column=7).value = address_field.get()

    wb.save(file_path)

    messagebox.showinfo("Success", "Data saved successfully!")

    name_field.focus_set()
    clear()


# main GUI
if __name__ == "__main__":

    root = Tk()
    root.configure(background='light green')
    root.title("Registration Form")
    root.geometry("600x400")

    excel()

    # labels
    Label(root, text="Form", bg="light green", font=("Arial", 16)).grid(row=0, column=1)
    Label(root, text="Name", bg="light green").grid(row=1, column=0)
    Label(root, text="Course", bg="light green").grid(row=2, column=0)
    Label(root, text="Semester", bg="light green").grid(row=3, column=0)
    Label(root, text="Form No.", bg="light green").grid(row=4, column=0)
    Label(root, text="Contact No.", bg="light green").grid(row=5, column=0)
    Label(root, text="Email ID", bg="light green").grid(row=6, column=0)
    Label(root, text="Address", bg="light green").grid(row=7, column=0)

    # entry fields
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    # bindings
    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)

    # grid placement
    name_field.grid(row=1, column=1, ipadx=100)
    course_field.grid(row=2, column=1, ipadx=100)
    sem_field.grid(row=3, column=1, ipadx=100)
    form_no_field.grid(row=4, column=1, ipadx=100)
    contact_no_field.grid(row=5, column=1, ipadx=100)
    email_id_field.grid(row=6, column=1, ipadx=100)
    address_field.grid(row=7, column=1, ipadx=100)

    # submit button
    Button(root, text="Submit", fg="black", bg="red",
           command=insert).grid(row=8, column=1)

    root.mainloop()